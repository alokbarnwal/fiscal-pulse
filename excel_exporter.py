"""
Fiscal Pulse — Office Excel Exporter (v2)
Layout: Group | Indicator | BE (FY1) | Apr | May | ... | Mar | BE (FY2) | Apr | May | ...
BE appears only ONCE per FY (annual figure), followed by monthly Actuals columns.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

from config import (
    INDICATOR_IDS, INDICATOR_NAMES, INDICATOR_GROUP,
    DQ_OK, DQ_MANUAL, DQ_PARSE_ERROR, DQ_PDF_ERROR, DQ_MISSING,
    MONTH_NAMES, state_to_filename
)

C_DARK   = "1F4E79"
C_MED    = "2E75B6"
C_LIGHT  = "D6E4F0"
C_ACCENT = "BDD7EE"
C_GREY   = "F5F5F5"
C_WHITE  = "FFFFFF"

GROUP_COLORS = {
    "Receipts":        ("DEEAF1", "1F4E79"),
    "Expenditure":     ("FBE5D6", "843C0C"),
    "Fiscal Position": ("E2EFDA", "375623"),
}

DQ_COLORS = {
    DQ_OK:          ("E2EFDA", "375623"),
    DQ_MANUAL:      ("D9E1F2", "203864"),
    DQ_PARSE_ERROR: ("FFF2CC", "7F6000"),
    DQ_PDF_ERROR:   ("FCE4D6", "843C0C"),
    DQ_MISSING:     ("EDEDED", "595959"),
}

MONTH_SHORT = {
    1:"Apr",2:"May",3:"Jun",4:"Jul",5:"Aug",6:"Sep",
    7:"Oct",8:"Nov",9:"Dec",10:"Jan",11:"Feb",12:"Mar"
}

def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(border_style="medium", color="1F4E79")
    return Border(left=s, right=s, top=s, bottom=s)

def fy_sep_border():
    thick = Side(border_style="medium", color="1F4E79")
    thin  = Side(border_style="thin",   color="BFBFBF")
    return Border(left=thin, right=thick, top=thin, bottom=thin)

def fmt_cell(ws, r, c, val=None, bold=False, fg="000000", bg=None,
             num_fmt=None, halign="center", wrap=True, size=9, italic=False,
             border=None, sep_right=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="Arial", bold=bold, size=size, color=fg, italic=italic)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt
    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    cell.border = border if border else (fy_sep_border() if sep_right else thin_border())
    return cell


def add_summary_sheet(wb, df):
    ws = wb.active
    ws.title = "📊 Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Fiscal Pulse — CAG Monthly Key Indicators"
    c.font  = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    s = ws["A2"]
    s.value = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Source: CAG India  |  Unit: ₹ Crore"
    s.font  = Font(name="Arial", size=9, color="FFFFFF", italic=True)
    s.fill  = PatternFill("solid", fgColor=C_MED)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = ["State", "Category", "Latest FY", "Latest Month", "Records", "Completeness"]
    for c, h in enumerate(headers, 1):
        fmt_cell(ws, 3, c, h, bold=True, fg="FFFFFF", bg=C_DARK, size=9)
    ws.row_dimensions[3].height = 20

    for r, state in enumerate(sorted(df["state"].unique()), start=4):
        sdf = df[df["state"] == state]
        bg = C_GREY if r % 2 == 0 else None
        latest_fy    = sdf["fy"].max()
        latest_month = sdf[sdf["fy"] == latest_fy]["month_name"].iloc[-1] if len(sdf) else "—"
        category     = sdf["category"].iloc[0] if "category" in sdf.columns else "—"
        total        = len(sdf)
        ok_count     = len(sdf[sdf["data_quality"] == DQ_OK])
        completeness = f"{ok_count/total*100:.1f}%" if total else "0%"

        fmt_cell(ws, r, 1, state,        bold=True, halign="left",   bg=bg, size=9)
        fmt_cell(ws, r, 2, category,                halign="center", bg=bg, size=9)
        fmt_cell(ws, r, 3, latest_fy,               halign="center", bg=bg, size=9)
        fmt_cell(ws, r, 4, latest_month,             halign="center", bg=bg, size=9)
        fmt_cell(ws, r, 5, total,        num_fmt="#,##0", bg=bg, size=9)
        fmt_cell(ws, r, 6, completeness, halign="center", bg=bg, size=9)
        ws.row_dimensions[r].height = 16

    for i, w in enumerate([22,12,12,14,10,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def add_state_sheet(wb, df, state_name):
    sdf = df[df["state"] == state_name].copy()
    if sdf.empty:
        return

    ws = wb.create_sheet(title=f"📍{state_to_filename(state_name)}")
    ws.sheet_view.showGridLines = False

    fys = sorted(sdf["fy"].unique())
    FIXED = 2   # Group + Indicator

    # Build col map: each FY = 1 BE col + 12 month cols = 13 cols
    col_map = {}
    col = FIXED + 1
    for fy in fys:
        col_map[(fy, "BE")] = col; col += 1
        for m in range(1, 13):
            col_map[(fy, m)] = col; col += 1
    total_cols = col - 1

    # Row 1 & 2 headers
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # Fixed col headers (merged rows 1-2)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    fmt_cell(ws, 1, 1, "Group",     bold=True, fg="FFFFFF", bg=C_DARK, size=9)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    fmt_cell(ws, 1, 2, "Indicator", bold=True, fg="FFFFFF", bg=C_DARK, size=9, halign="left")

    for fy in fys:
        is_last_fy = (fy == fys[-1])
        start_c = col_map[(fy, "BE")]
        end_c   = col_map[(fy, 12)]
        ws.merge_cells(start_row=1, start_column=start_c, end_row=1, end_column=end_c)
        fc = ws.cell(row=1, column=start_c,
                     value=f"FY {fy}  |  {state_name}  |  ₹ Crore")
        fc.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        fc.fill      = PatternFill("solid", fgColor=C_DARK)
        fc.alignment = Alignment(horizontal="center", vertical="center")
        fc.border    = medium_border()

        # BE sub-header
        fmt_cell(ws, 2, col_map[(fy,"BE")], "BE\n(Annual)",
                 bold=True, fg="FFFFFF", bg=C_MED, size=8, wrap=True)
        # Month sub-headers
        for m in range(1, 13):
            sep = (m == 12) and not is_last_fy
            fmt_cell(ws, 2, col_map[(fy,m)], MONTH_SHORT[m],
                     bold=True, fg="FFFFFF", bg=C_MED, size=8, sep_right=sep)

    # Lookup tables
    be_lookup = {}
    for fy in fys:
        fy_df = sdf[sdf["fy"] == fy]
        for ind_id in INDICATOR_IDS:
            ind_df = fy_df[fy_df["indicator_id"] == ind_id]
            if not ind_df.empty:
                be_vals = ind_df["be"].dropna()
                be_lookup[(ind_id, fy)] = be_vals.max() if not be_vals.empty else None
            else:
                be_lookup[(ind_id, fy)] = None

    actuals_lookup = {}
    dq_lookup      = {}
    for _, row in sdf.iterrows():
        key = (row["indicator_id"], row["fy"], int(row["fy_month_order"]))
        actuals_lookup[key] = row["actuals_cumulative"]
        dq_lookup[key]      = row["data_quality"]

    # Data rows
    row_num    = 3
    prev_group = None

    for ind_id in INDICATOR_IDS:
        group = INDICATOR_GROUP.get(ind_id, "")
        name  = INDICATOR_NAMES.get(ind_id, ind_id)
        is_total = ("total" in ind_id or ind_id in (
            "revenue_receipts","revenue_surplus_deficit","fiscal_deficit","primary_deficit"))

        # Group header
        if group != prev_group:
            g_bg, g_fg = GROUP_COLORS.get(group, (C_LIGHT, C_DARK))
            ws.merge_cells(start_row=row_num, start_column=1,
                           end_row=row_num, end_column=total_cols)
            gc = ws.cell(row=row_num, column=1, value=group.upper())
            gc.font      = Font(name="Arial", bold=True, size=9, color=g_fg)
            gc.fill      = PatternFill("solid", fgColor=g_bg)
            gc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            gc.border    = thin_border()
            ws.row_dimensions[row_num].height = 16
            row_num   += 1
            prev_group = group

        row_bg = C_LIGHT if is_total else None
        row_fg = C_DARK  if is_total else "000000"

        fmt_cell(ws, row_num, 1, group, fg="999999", size=8, halign="left", bg=row_bg)
        fmt_cell(ws, row_num, 2, name,  bold=is_total, fg=row_fg, bg=row_bg,
                 halign="left", size=9)

        for fy in fys:
            is_last_fy = (fy == fys[-1])
            be_val = be_lookup.get((ind_id, fy))
            fmt_cell(ws, row_num, col_map[(fy,"BE")],
                     be_val if be_val is not None else "—",
                     bold=True, fg=C_DARK, bg=C_ACCENT,
                     num_fmt="#,##0" if isinstance(be_val, float) else None, size=9)

            for m in range(1, 13):
                sep    = (m == 12) and not is_last_fy
                key    = (ind_id, fy, m)
                actual = actuals_lookup.get(key)
                dq     = dq_lookup.get(key, DQ_MISSING)

                if actual is not None:
                    dq_bg, dq_fg = DQ_COLORS.get(dq, (C_WHITE, "000000"))
                    fmt_cell(ws, row_num, col_map[(fy,m)], actual,
                             fg=dq_fg, bg=dq_bg, num_fmt="#,##0.##",
                             size=9, sep_right=sep)
                else:
                    fmt_cell(ws, row_num, col_map[(fy,m)], None,
                             fg="BBBBBB", bg=C_GREY, size=9, sep_right=sep)

        ws.row_dimensions[row_num].height = 15
        row_num += 1

    # Legend
    row_num += 1
    fmt_cell(ws, row_num, 1, "Data Quality:", bold=True, fg=C_DARK,
             size=8, halign="left")
    for lc, (dq_label, (dq_bg, dq_fg)) in enumerate(DQ_COLORS.items(), start=2):
        fmt_cell(ws, row_num, lc, dq_label, bold=True, fg=dq_fg, bg=dq_bg, size=8)

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    col = FIXED + 1
    for fy in fys:
        ws.column_dimensions[get_column_letter(col)].width = 10; col += 1
        for _ in range(12):
            ws.column_dimensions[get_column_letter(col)].width = 7; col += 1

    ws.freeze_panes = "C3"


def add_dq_sheet(wb, df):
    ws = wb.create_sheet(title="🔍 Data Quality")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Data Quality Log — Issues Only (PARSE_ERROR / PDF_ERROR / MISSING)"
    t.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    t.fill  = PatternFill("solid", fgColor=C_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c, h in enumerate(["State","FY","Month","Indicator","Data Quality","BE","Actuals","PDF URL"], 1):
        fmt_cell(ws, 2, c, h, bold=True, fg="FFFFFF", bg=C_MED, size=9)
    ws.row_dimensions[2].height = 18

    issues = df[df["data_quality"] != DQ_OK].sort_values(
        ["state","fy","fy_month_order","indicator_id"])

    for r, (_, row) in enumerate(issues.iterrows(), start=3):
        dq_bg, dq_fg = DQ_COLORS.get(row["data_quality"], (C_WHITE,"000000"))
        bg = C_GREY if r % 2 == 0 else None
        fmt_cell(ws, r, 1, row["state"],              halign="left",   bg=bg, size=9)
        fmt_cell(ws, r, 2, row["fy"],                 halign="center", bg=bg, size=9)
        fmt_cell(ws, r, 3, row["month_name"],         halign="center", bg=bg, size=9)
        fmt_cell(ws, r, 4, row["indicator_name"],     halign="left",   bg=bg, size=9)
        dq_c = ws.cell(row=r, column=5, value=row["data_quality"])
        dq_c.font      = Font(name="Arial", size=9, bold=True, color=dq_fg)
        dq_c.fill      = PatternFill("solid", fgColor=dq_bg)
        dq_c.border    = thin_border()
        dq_c.alignment = Alignment(horizontal="center", vertical="center")
        fmt_cell(ws, r, 6, row["be"],                 num_fmt="#,##0", bg=bg, size=9)
        fmt_cell(ws, r, 7, row["actuals_cumulative"], num_fmt="#,##0", bg=bg, size=9)
        url_c = ws.cell(row=r, column=8, value=row["pdf_url"])
        url_c.font      = Font(name="Arial", size=8, color="0000EE", underline="single")
        url_c.border    = thin_border()
        url_c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 16

    for i, w in enumerate([22,10,12,30,15,14,14,60], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def generate_office_excel(df, output_path):
    """Generate formatted Excel. Called by pipeline.py automatically."""
    wb = openpyxl.Workbook()
    add_summary_sheet(wb, df)
    for state in sorted(df["state"].unique()):
        add_state_sheet(wb, df, state)
    add_dq_sheet(wb, df)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel saved → {output_path}")


if __name__ == "__main__":
    from config import MASTER_CSV, OFFICE_EXCEL
    print("Loading master.csv...")
    df = pd.read_csv(MASTER_CSV)
    print(f"  Rows: {len(df)} | States: {df['state'].nunique()} | FYs: {list(df['fy'].unique())}")
    generate_office_excel(df, OFFICE_EXCEL)
